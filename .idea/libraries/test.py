#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @File  : test.py
# @Author: xuan
# @Date  : 2019-03-22
# @Desc  : for test
import pandas as pd
from pandas import DataFrame,Series
import numpy as np
import tools
import csv
import chardet
import os
import time
from openpyxl import load_workbook
from openpyxl import Workbook


inp = [{'c1':10, 'c2':100}, {'c1':11,'c2':110}, {'c1':12,'c2':120}]
df = pd.DataFrame(inp)
print(df)
print(list(df['c1'].values))
if 12 in list(df['c1'].values):
    print('true')
else:
    print('false')

# def check_file_charset(file):
#     with open(file,'rb') as f:
#         return chardet.detect(f.read())
# print(check_file_charset('/users/xuan/desktop/SNA/data/edge_with_match_info.csv'))
# def decompose(nextState):
#     nextComList = []
#     tempList = nextState.split(' ')
#     event = tempList[0]
#     nextComList = tempList[1].split(',')
#     return event,nextComList
#
# string = 'splittingTo '+'C11,C22,C33,C44'
# event,nextComList = decompose(string)
# print(event)
# print(nextComList)
# while len(nextComList)!=2:
#     nextComList.pop()
# print(nextComList)

# def test():
#     return 1,2
# print(test()[0])
# p = '/users/xuan/desktop/SNA/data/Original/'
# wb_in = load_workbook(p+'20190410.xlsx')
# ws_in = wb_in.get_sheet_by_name('20190410')
# print(ws_in)
# team_info = {}
# edges_count = 0
# for i in range(2, ws_in.max_row):
#     team_id = ws_in.cell(i, 2).value
#     if team_info.__contains__(team_id) == False:
#         team_info[team_id] = []
#     team_info[team_id].append(ws_in.cell(i, 1).value)
#
# # 输出'NoneType' object has no attribute
# wb_out = Workbook()
# ws_out = wb_out.active
# ws_out.append(['source', 'target', 'team_id'])
# for key in team_info.keys():
#     team_members = team_info[key]
#     team_len = len(team_info[key])
#     # print(team_len, team_members)
#     handshake_count = 0
#     if (team_len > 1)&(team_len < 50):
#         for i in range(0, team_len):
#             for j in range(i + 1, team_len):
#                 # print(i, j, [team_members[i], team_members[j], key])
#                 ws_out.append([team_members[i], team_members[j], key])
#                 handshake_count += 1
#     edges_count += handshake_count
#     print('team_id:', key, ' team_len:', team_len, ' handshakes:', handshake_count)
# print("共计形成边：",edges_count)
# wb_out.save(p+'edegs20190411.xlsx')
# def transfer_event_about_match(f):
#     df = pd.read_csv(f,low_memory=False,encoding='UTF-8')
#     result_transfer = DataFrame({'slot':[],'event':[],'match_type':[],'count':[]})
#     all_slots = Series(df.slot).unique()
#     for slot in all_slots:
#         data_in_spec_slot = df.loc[df['slot']==slot]
#         event_in_spec_slot = Series(data_in_spec_slot.event).unique()
#         for event in event_in_spec_slot:
#             spec_event_data = data_in_spec_slot.loc[data_in_spec_slot['event']==event]
#             m1 = spec_event_data['m1'].sum()
#             m2 = spec_event_data['m2'].sum()
#             m3 = spec_event_data['m3'].sum()
#             result_transfer = result_transfer.append(
#                 DataFrame({'slot': [slot], 'event': [event], 'match_type': ['m1'], 'count': [m1]}))
#             result_transfer = result_transfer.append(
#                 DataFrame({'slot': [slot], 'event': [event], 'match_type': ['m2'], 'count': [m2]}))
#             result_transfer = result_transfer.append(
#                 DataFrame({'slot': [slot], 'event': [event], 'match_type': ['m3'], 'count': [m3]}))
#             print(result_transfer)
#     result_transfer.to_csv('/users/xuan/desktop/SNA/data/result_transfer.csv')
#print(tools.timeTrans(1556152269,"%Y-%m-%d %H:%M:%S"))
#print(tools.timeTrans(1556221190,"%Y-%m-%d %H:%M:%S"))