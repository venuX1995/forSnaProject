from openpyxl import load_workbook
from openpyxl import Workbook

# 读取
p = '/users/xuan/desktop/SNA/data/Original/'
wb_in = load_workbook(p+'20190410.xlsx')
ws_in = wb_in.get_sheet_by_name('20190410')
print(ws_in)
team_info = {}
edges_count = 0
for i in range(2, ws_in.max_row):
    team_id = ws_in.cell(i, 2).value
    if team_info.__contains__(team_id) == False:
        team_info[team_id] = []
    team_info[team_id].append(ws_in.cell(i, 1).value)

# 输出'NoneType' object has no attribute
wb_out = Workbook()
ws_out = wb_out.active
ws_out.append(['source', 'target', 'team_id'])
for key in team_info.keys():
    team_members = team_info[key]
    team_len = len(team_info[key])
    # print(team_len, team_members)
    handshake_count = 0
    if (team_len > 1)&(team_len < 50):
        for i in range(0, team_len):
            for j in range(i + 1, team_len):
                # print(i, j, [team_members[i], team_members[j], key])
                ws_out.append([team_members[i], team_members[j], key])
                handshake_count += 1
    edges_count += handshake_count
    print('team_id:', key, ' team_len:', team_len, ' handshakes:', handshake_count)
print("共计形成边：",edges_count)
wb_out.save(p+'edegs20190411.xlsx')